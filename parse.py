''' Python Library Dependencies 
numpy
xlrd
openpyxl
'''

import numpy as np
import openpyxl as op

#
#Parse player ratings and add to Excel spreadsheets
#

#Add player ratings and positions
wb = op.load_workbook(filename = 'nba2kplayerratings.xlsx')
sheet = wb['Sheet1']

#Add 2K13 player ratings to column C of excel sheet
initialtuple = ()
playerratings = dict((y.lower(), x) for x, y in initialtuple)
playersmatched = []
for row in range(2, sheet.max_row):
	playername = sheet['A'+str(row)].value.lower()
	if playername in playerratings: #If we have data on this player
		sheet['C'+str(row)].value = int(playerratings[playername])
		playersmatched.append(playername)
	elif sheet['D'+str(row)].value=='-': #If this player joined the league after 2014
		sheet['C'+str(row)].value = '-'
	else:
		print playername

#Add 2K14 player ratings to column D of excel sheet
initialtuple = ()
playerratings = dict((y.lower(), x) for x, y in initialtuple)
playersmatched = []
for row in range(2, sheet.max_row):
	playername = sheet['A'+str(row)].value.lower()
	if playername in playerratings: #If we have data on this player
		sheet['D'+str(row)].value = int(playerratings[playername])
		playersmatched.append(playername)
	elif sheet['E'+str(row)].value=='-': #If this player joined the league after 2014
		sheet['D'+str(row)].value = '-'
	else:
		print playername

playersnotmatched = []
for p in playerratings:
	if p not in playersmatched: #This player isn't in the Excel sheet
		rowno = sheet.max_row+1
		sheet['A'+str(rowno)].value = p.title()
		sheet['D'+str(rowno)].value = playerratings[p]

#Add positions to 2K13 players
for row in range(2, sheet.max_row):
	playername = sheet['A'+str(row)].value
	val = ""
	if playername in pg:
		val = "PG"
	elif playername in sg:
		val = "SG"
	elif playername in sf:
		val = "SF"
	elif playername in pf:
		val = "PF"
	elif playername in c:
		val = "C"
	else:
		continue
	sheet['G'+str(row)].value = val

for row in range(2, sheet.max_row):
	val = sheet['G'+str(row)].value
	if not (val=="PG" or val=="SG" or val=="SF" or val=="PF" or val=="C"):
		print "#"+str(row)+" - "+sheet["A"+str(row)].value

wb.save('nba2kplayerratings.xlsx')